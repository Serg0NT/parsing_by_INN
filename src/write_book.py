from pathlib import Path
import datetime

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from parsing import get_data

_path_to_results_dir = str(Path(__file__).parents[1]) + "\\" + "results" + "\\"

# заголовки файла с результатами
headers = ['Статус', 'ИНН', 'Наименование', 'Полное наименование',
           'Род деятельности', 'Адрес', 'Выручка', 'Прирост']

# абсолютный путь до файла с результатами парсинга
workbook_name = _path_to_results_dir + 'result_file.xlsx'

# задает имя листа как текущая дата и время
name_sheet = datetime.datetime.today().strftime('%d-%m-%Y %H.%M')

# пытаемся открыть существующий файл с результатами.
# если его нет, создаем новый
try:
    wb = load_workbook(workbook_name)
    page = wb.create_sheet(name_sheet, 0)
except FileNotFoundError:
    wb = Workbook()
    page = wb.active
    page.title = name_sheet

wb.active = 0  # первый лист делаем активным
page.append(headers)  # записываем заголовки столбцов
wb.save(filename=workbook_name)  # сохраняем файл


# добавляем данные в таблицу
def write_data():
    """
    Сохраняет полученные данные в эксель
    """
    count = 2

    for info in get_data():
        print(info)
        wb1 = load_workbook(workbook_name)
        page1 = wb1[name_sheet]
        page1.append(info)
        wb1.save(filename=workbook_name)

        print(f"Обработана строка № {count}")
        count += 1
    print(f"Это было не легко, но мы справились!!!")
