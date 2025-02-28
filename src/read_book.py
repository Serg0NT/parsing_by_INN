import os
from openpyxl import load_workbook
from pathlib import Path


def _get_input_file() -> str:
    """
    ищет путь до папки input_data и выводит по очереди файлы ексель
    :return: str
    """
    path_to_input_dir = str(Path(__file__).parents[1]) + "\\" + "input_data"
    for file in os.listdir(path_to_input_dir):
        if file.endswith(".xlsx") or file.endswith(".xls"):
            yield path_to_input_dir + "\\" + file


def import_data() -> str:
    """
    ищет номер столбца с заголовком ИНН и возвращает по очереди все ИНН
    :return: str
    """
    for path in _get_input_file():
        dataframe = load_workbook(path)
        dataframe1 = dataframe.active
        # Ищем номер нужного столбца
        nmb_column_inn = 0
        for row in range(0, 1):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                if col[row].value == "ИНН":
                    nmb_column_inn = nmb_column_inn
                    break
                nmb_column_inn += 1

        # считываем данные из столбца инн
        for row in dataframe1.iter_rows(2, dataframe1.max_row):
            yield row[nmb_column_inn].value
